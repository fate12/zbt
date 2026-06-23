#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把「结构化 活动信息.xlsx」转成适合阿里云百炼知识库的 Markdown。
关键修复：把单元格 hyperlink.target（真实 URL）内联进文本，
否则百炼解析 xlsx 时只会拿到显示文字，URL 丢失。
每行 = 一个 Markdown 块，块之间用「===ACTIVITY===」分隔，
便于在百炼里用「自定义分段符」精准切片（一行一 chunk）。
"""
import sys
import openpyxl

SRC = "/Users/liuyuanyuan/Downloads/结构化 活动信息.xlsx"
DST = "/Users/liuyuanyuan/Downloads/结构化活动信息_知识库.md"

# (列序号, 字段名)
FIELDS = [
    (1, "活动名称"),
    (2, "活动分类"),
    (3, "所属板块"),
    (4, "活动标签"),
    (5, "报名对象"),
    (6, "报名要求描述"),
    (7, "详细活动要求"),
    (8, "报名要求图片"),
    (10, "报名链接1"),
    (11, "报名链接2"),
    (12, "备注说明"),
]
# 这些列把 hyperlink 真实 URL 内联进来；其它列只取显示文字
INLINE_LINK_COLS = {10, 11}
SEP = "\n\n===ACTIVITY===\n\n"


def cell_text(ws, r, c):
    cell = ws.cell(row=r, column=c)
    text = "" if cell.value is None else str(cell.value).strip()
    link = cell.hyperlink.target if cell.hyperlink else None
    if link and c in INLINE_LINK_COLS:
        # 文字本身已经是 URL 就只输出 URL，否则「文字(URL)」
        if text and text != link:
            return f"{text}({link})"
        return link
    return text


def build_header_map(ws):
    """读第一行表头，建立「字段名 -> 列号」映射，用于按名称定位列号不固定的可选列。"""
    m = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            m[str(v).strip()] = c
    return m


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["数据表"]
    header = build_header_map(ws)
    # 「推荐频次」列号在 Excel 中不固定，按表头名动态查找；缺失则不输出该字段
    freq_col = header.get("推荐频次")
    blocks = []
    for r in range(2, ws.max_row + 1):
        name = cell_text(ws, r, 1)
        if not name:
            continue
        lines = []
        for c, label in FIELDS:
            val = cell_text(ws, r, c)
            lines.append(f"{label}：{val}" if val else f"{label}：")
        if freq_col is not None:
            freq_val = cell_text(ws, r, freq_col)
            lines.append(f"推荐频次：{freq_val}" if freq_val else "推荐频次：")
        blocks.append("\n".join(lines))
    out = SEP.join(blocks)
    with open(DST, "w", encoding="utf-8") as f:
        f.write(out + "\n")
    print(f"done: {len(blocks)} 条活动 -> {DST}")


if __name__ == "__main__":
    main()
